import json
from django.shortcuts import render, get_object_or_404
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.views.generic import TemplateView
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
from django.db.models import Sum, Count, F, Q, Max
from django.db import models
from django.utils import timezone
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import Order, OrderItem, OrderStage, OrderDefect
from .serializers import OrderSerializer, OrderItemSerializer, OrderStageConfirmSerializer, OrderStageSerializer
from apps.employee_tasks.models import EmployeeTask
from apps.employees.models import User

# Create your views here.

class OrderViewSet(viewsets.ModelViewSet):
	queryset = Order.objects.select_related('client', 'workshop', 'product').prefetch_related('items__product', 'stages__workshop', 'order_defects__workshop').all().order_by('-created_at')
	serializer_class = OrderSerializer
	permission_classes = [permissions.IsAuthenticated]
	
	@action(detail=False, methods=['get'])
	def by_workshop(self, request):
		"""Получает заказы с разделением по цехам"""
		workshop_id = request.query_params.get('workshop_id')
		
		if workshop_id:
			# Фильтруем заказы по конкретному цеху
			queryset = Order.objects.filter(
				stages__workshop_id=workshop_id,
				stages__status__in=['in_progress', 'partial']
			).distinct().order_by('-created_at')
		else:
			# Получаем все заказы с группировкой по цехам
			queryset = Order.objects.filter(
				stages__status__in=['in_progress', 'partial']
			).distinct().order_by('-created_at')
		
		# Добавляем информацию о разделении по цехам
		orders_data = []
		for order in queryset:
			order_data = OrderSerializer(order).data
			
			# Добавляем информацию о цехах
			workshops_info = {}
			for stage in order.stages.filter(status__in=['in_progress', 'partial']):
				workshop_name = stage.workshop.name if stage.workshop else 'Не указан'
				workshop_type = 'Стеклянные товары' if stage.parallel_group == 1 else 'Обычные товары'
				
				if workshop_name not in workshops_info:
					workshops_info[workshop_name] = {
						'type': workshop_type,
						'quantity': 0,
						'operation': stage.operation
					}
				workshops_info[workshop_name]['quantity'] += stage.plan_quantity
			
			order_data['workshops_info'] = workshops_info
			order_data['has_glass_items'] = order.has_glass_items
			order_data['has_regular_items'] = order.regular_items
			
			orders_data.append(order_data)
		
		return Response(orders_data)
	
	def update(self, request, *args, **kwargs):
		instance = self.get_object()
		data = request.data.copy()
		
		# Обрабатываем позиции заказа отдельно
		items_data = data.pop('items_data', None)
		
		# Обновляем основные поля заказа
		serializer = self.get_serializer(instance, data=data, partial=True)
		serializer.is_valid(raise_exception=True)
		self.perform_update(serializer)
		
		# Если переданы новые позиции, обновляем их
		if items_data is not None:
			# Удаляем старые позиции
			instance.items.all().delete()
			
			# Создаем новые позиции
			for item_data in items_data:
				product_id = item_data.get('product_id')
				quantity = item_data.get('quantity', 1)
				size = item_data.get('size', '')
				color = item_data.get('color', '')
				glass_type = item_data.get('glass_type', '')
				paint_type = item_data.get('paint_type', '')
				paint_color = item_data.get('paint_color', '')
				cnc_specs = item_data.get('cnc_specs', '')
				cutting_specs = item_data.get('cutting_specs', '')
				packaging_notes = item_data.get('packaging_notes', '')
				
				# Проверяем что product_id существует и валиден
				if not product_id:
					continue
					
				from apps.products.models import Product
				try:
					product = Product.objects.get(pk=product_id)
				except Product.DoesNotExist:
					continue  # Пропускаем несуществующие товары
				
				OrderItem.objects.create(
					order=instance,
					product=product,
					quantity=quantity,
					size=size,
					color=color,
					glass_type=glass_type,
					paint_type=paint_type,
					paint_color=paint_color,
					cnc_specs=cnc_specs,
					cutting_specs=cutting_specs,
					packaging_notes=packaging_notes
				)
			
			# Удаляем старые этапы и создаем новые
			instance.stages.all().delete()
			from .models import create_order_stages
			try:
				create_order_stages(instance)
			except Exception as e:
				print(f"Warning: Error creating order stages: {e}")
				# Продолжаем выполнение, этапы не критичны для обновления заказа
		
		# Возвращаем обновленный заказ
		return Response(OrderSerializer(instance).data)

class OrderCreateAPIView(APIView):
	permission_classes = [permissions.IsAuthenticated]
	@method_decorator(csrf_exempt)
	def post(self, request):
		try:
			data = request.data
			name = data.get('name')
			client_id = data.get('client_id')
			items_data = data.get('items_data', [])
			
			if not name or not client_id or not items_data:
				return Response({
					'error': 'Необходимо указать название заказа, клиента и товары'
				}, status=status.HTTP_400_BAD_REQUEST)
			
			from apps.clients.models import Client
			client = get_object_or_404(Client, pk=client_id)
			
			# Создаем заказ
			order = Order.objects.create(
				name=name,
				client=client,
				status='production'
			)
			
			# Создаем позиции заказа
			created_items = []
			for item_data in items_data:
				product_id = item_data.get('product_id')
				quantity = item_data.get('quantity', 1)
				size = item_data.get('size', '')
				color = item_data.get('color', '')
				glass_type = item_data.get('glass_type', '')
				paint_type = item_data.get('paint_type', '')
				paint_color = item_data.get('paint_color', '')
				cnc_specs = item_data.get('cnc_specs', '')
				cutting_specs = item_data.get('cutting_specs', '')
				packaging_notes = item_data.get('packaging_notes', '')
				
				from apps.products.models import Product
				product = get_object_or_404(Product, pk=product_id)
				
				item = OrderItem.objects.create(
					order=order,
					product=product,
					quantity=quantity,
					size=size,
					color=color,
					glass_type=glass_type,
					paint_type=paint_type,
					paint_color=paint_color,
					cnc_specs=cnc_specs,
					cutting_specs=cutting_specs,
					packaging_notes=packaging_notes
				)
				created_items.append(item)
			
			# Автоматически создаем этапы заказа после создания всех позиций
			from .models import create_order_stages
			create_order_stages(order)
			
			# Возвращаем созданный заказ с полной информацией
			return Response(OrderSerializer(order).data, status=status.HTTP_201_CREATED)
			
		except Exception as e:
			print('ORDER CREATE ERROR:', str(e))
			return Response({
				'error': f'Ошибка создания заказа: {str(e)}'
			}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class OrderPageView(View):
	def get(self, request):
		user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
		is_mobile = any(m in user_agent for m in ['android', 'iphone', 'ipad', 'mobile'])
		template = 'orders_mobile.html' if is_mobile else 'orders.html'
		show_create = request.GET.get('create') == 'True' or request.GET.get('create') == 'true' or request.GET.get('create') == '1'
		
		# Получаем статистику по заказам с разделением на стеклянные и обычные
		from apps.orders.models import Order
		from apps.operations.workshops.models import Workshop
		
		# Статистика по цехам
		workshops_stats = {}
		try:
			workshop_1 = Workshop.objects.get(pk=1)  # Обычные товары
			workshop_2 = Workshop.objects.get(pk=2)  # Стеклянные товары
			
			# Заказы в цехе 1 (обычные товары)
			regular_orders = Order.objects.filter(
				stages__workshop=workshop_1,
				stages__status__in=['in_progress', 'partial']
			).distinct().count()
			
			# Заказы в цехе 2 (стеклянные товары)
			glass_orders = Order.objects.filter(
				stages__workshop=workshop_2,
				stages__status__in=['in_progress', 'partial']
			).distinct().count()
			
			workshops_stats = {
				'workshop_1': {
					'name': workshop_1.name,
					'orders_count': regular_orders,
					'type': 'Обычные товары'
				},
				'workshop_2': {
					'name': workshop_2.name,
					'orders_count': glass_orders,
					'type': 'Стеклянные товары'
				}
			}
		except Workshop.DoesNotExist:
			workshops_stats = {}
		
		context = {
			'show_create': show_create,
			'workshops_stats': workshops_stats
		}
		return render(request, template, context)

class OrderStageConfirmAPIView(APIView):
	permission_classes = [permissions.IsAuthenticated]
	def patch(self, request, stage_id):
		stage = get_object_or_404(OrderStage, pk=stage_id)
		serializer = OrderStageConfirmSerializer(data=request.data)
		if serializer.is_valid():
			completed_qty = serializer.validated_data['completed_quantity']
			stage.confirm_stage(completed_qty)
			return Response({'status': 'ok', 'stage': stage.id, 'completed_quantity': completed_qty})
		return Response(serializer.errors, status=400)

class OrderStageTransferAPIView(APIView):
	permission_classes = [permissions.IsAuthenticated]
	def post(self, request, stage_id):
		stage = get_object_or_404(OrderStage, pk=stage_id)
		# Поддержка явного выбора цеха и количества, без жёсткого workflow
		target_workshop_id = request.data.get('target_workshop_id')
		completed_qty = request.data.get('completed_quantity')
		try:
			completed_qty = int(completed_qty) if completed_qty is not None else stage.plan_quantity
		except (TypeError, ValueError):
			completed_qty = stage.plan_quantity
		if completed_qty < 0:
			completed_qty = 0
		if completed_qty > stage.plan_quantity:
			completed_qty = stage.plan_quantity
		
		# Ограничения и логика для стеклянных изделий: только цеха 2 и 12; при переводе в 12 — агрегируем
		is_glass = bool(getattr(getattr(getattr(stage, 'order_item', None), 'product', None), 'is_glass', False))
		current_workshop_id = getattr(getattr(stage, 'workshop', None), 'id', None)
		
		if target_workshop_id:
			try:
				target_workshop_id = int(target_workshop_id)
			except (TypeError, ValueError):
				return Response({'error': 'target_workshop_id must be an integer'}, status=400)
			
			if is_glass:
				# Допускаем только перемещения между цехами 2 и 12
				if current_workshop_id not in (2, 12):
					return Response({'error': 'Glass items may only be processed in workshops 2 and 12'}, status=400)
				if target_workshop_id not in (2, 12):
					return Response({'error': 'Glass items can only be transferred to workshop 2 or 12'}, status=400)
				
				# Завершаем текущий этап на указанное количество (частично или полностью)
				stage.confirm_stage(completed_qty)
				from apps.operations.workshops.models import Workshop
				workshop = get_object_or_404(Workshop, pk=target_workshop_id)
				
				# Если переводим в 12 — агрегируем по заявке (order_item=NULL) в цехе 12
				if target_workshop_id == 12:
					# Ищем существующий агрегирующий этап по заказу без привязки к позиции в 12 цехе
					aggregate_stage = OrderStage.objects.filter(
						order=stage.order,
						order_item__isnull=True,
						workshop_id=12,
						stage_type='workshop',
						parallel_group=stage.parallel_group,
					).order_by('sequence').first()
					if aggregate_stage:
						aggregate_stage.plan_quantity += completed_qty
						aggregate_stage.status = 'in_progress'
						aggregate_stage.save(update_fields=['plan_quantity', 'status'])
					else:
						# Создаем новый агрегирующий этап
						OrderStage.objects.create(
							order=stage.order,
							order_item=None,
							sequence=(stage.sequence or 0) + 1,
							stage_type='workshop',
							workshop=workshop,
							operation=f"Сборка заказа (стекло) из: {stage.workshop.name if stage.workshop else ''}",
							plan_quantity=completed_qty,
							deadline=timezone.now().date(),
							status='in_progress',
							parallel_group=stage.parallel_group,
						)
					return Response({'status': 'ok', 'stage': stage.id, 'action': 'transferred', 'target_workshop_id': target_workshop_id, 'completed_quantity': completed_qty, 'aggregated': True})
				
				# Иначе (перевод в 2) — ведем обычным образом, но только внутри 2/12
				# Ищем следующий этап по заказу и позиции в выбранном цехе (без учёта sequence)
				next_stage = OrderStage.objects.filter(
					order=stage.order,
					order_item=stage.order_item,
					workshop_id=target_workshop_id,
					stage_type='workshop',
					parallel_group=stage.parallel_group,
				).order_by('sequence').first()
				if next_stage:
					next_stage.plan_quantity = completed_qty
					next_stage.status = 'in_progress'
					next_stage.save(update_fields=['plan_quantity', 'status'])
				else:
					OrderStage.objects.create(
						order=stage.order,
						order_item=stage.order_item,
						sequence=(stage.sequence or 0) + 1,
						stage_type='workshop',
						workshop=workshop,
						operation=f"Передано из: {stage.workshop.name if stage.workshop else ''}",
						plan_quantity=completed_qty,
						deadline=timezone.now().date(),
						status='in_progress',
						parallel_group=stage.parallel_group,
					)
				return Response({'status': 'ok', 'stage': stage.id, 'action': 'transferred', 'target_workshop_id': target_workshop_id, 'completed_quantity': completed_qty})
			
			# Нестеклянные — прежняя логика явного перевода
			# Завершаем текущий этап на указанное количество (частично или полностью)
			stage.confirm_stage(completed_qty)
			
			# АВТОМАТИЧЕСКОЕ УДАЛЕНИЕ СТЕКЛЯННЫХ ТОВАРОВ ПРИ ПЕРЕВОДЕ С ПРЕССА (ID5)
			if current_workshop_id == 5 and target_workshop_id > 5:
				# Удаляем все стеклянные товары из заказа при переводе с пресса
				glass_items = stage.order.items.filter(product__is_glass=True)
				if glass_items.exists():
					glass_count = glass_items.count()
					glass_items.delete()
					# Логируем удаление
					print(f"Удалено {glass_count} стеклянных товаров из заказа {stage.order.id} при переводе с пресса")
			
			# Явно создаём/активируем этап в выбранном цехе
			from apps.operations.workshops.models import Workshop
			workshop = get_object_or_404(Workshop, pk=target_workshop_id)
			# Ищем следующий этап по заказу и позиции в выбранном цехе (без учёта sequence)
			next_stage = OrderStage.objects.filter(
				order=stage.order,
				order_item=stage.order_item,
				workshop=workshop,
				stage_type='workshop',
				parallel_group=stage.parallel_group,
			).order_by('sequence').first()
			if next_stage:
				next_stage.plan_quantity = completed_qty
				next_stage.status = 'in_progress'
				next_stage.save()
			else:
				# Создаём новый этап в выбранном цехе
				OrderStage.objects.create(
					order=stage.order,
					order_item=stage.order_item,
					sequence=stage.sequence + 1,
					stage_type='workshop',
					workshop=workshop,
					operation=f"Передано из: {stage.workshop.name if stage.workshop else ''}",
					plan_quantity=completed_qty,
					deadline=timezone.now().date(),
					status='in_progress',
					parallel_group=stage.parallel_group,
				)
			return Response({'status': 'ok', 'stage': stage.id, 'action': 'transferred', 'target_workshop_id': int(target_workshop_id), 'completed_quantity': completed_qty})
		
		# Fallback: прежнее поведение — перевод по workflow всего плана
		stage.confirm_stage(stage.plan_quantity)
		return Response({'status': 'ok', 'stage': stage.id, 'action': 'transferred'})

class OrderStagePostponeAPIView(APIView):
	permission_classes = [permissions.IsAuthenticated]
	def post(self, request, stage_id):
		stage = get_object_or_404(OrderStage, pk=stage_id)
		# Переносим дедлайн на следующий рабочий день (просто +1 день)
		if stage.deadline:
			from datetime import timedelta
			stage.deadline = stage.deadline + timedelta(days=1)
			stage.save()
			return Response({'status': 'ok', 'stage': stage.id, 'new_deadline': stage.deadline, 'action': 'postponed'})
		return Response({'status': 'error', 'error': 'No deadline set'}, status=400)

class OrderStageNoTransferAPIView(APIView):
	permission_classes = [permissions.IsAuthenticated]
	def post(self, request, stage_id):
		stage = get_object_or_404(OrderStage, pk=stage_id)
		# Фиксируем этап как "не переводить" (например, статус waiting)
		stage.status = 'waiting'
		stage.save()
		return Response({'status': 'ok', 'stage': stage.id, 'action': 'no_transfer'})

class DashboardOverviewAPIView(APIView):
	permission_classes = [permissions.IsAuthenticated]
	def get(self, request):
		# Доход — сумма (цена продукта * количество) по всем позициям заявок
		total_income = OrderItem.objects.aggregate(total=Sum(F('product__price') * F('quantity')))['total'] or 0
		# Продажи — сумма quantity по всем позициям
		product_sales = OrderItem.objects.aggregate(total=Sum('quantity'))['total'] or 0
		# Брак — сумма quantity по всем OrderDefect + сумма defective_quantity по всем EmployeeTask
		order_defects_total = OrderDefect.objects.aggregate(total=Sum('quantity'))['total'] or 0
		employee_tasks_defects_total = EmployeeTask.objects.aggregate(total=Sum('defective_quantity'))['total'] or 0
		defective_products = order_defects_total + employee_tasks_defects_total
		# Сотрудники — всего
		total_employees = User.objects.count()
		return Response({
			'total_income': total_income,
			'product_sales': product_sales,
			'defective_products': defective_products,
			'total_employees': total_employees,
			'user_name': request.user.get_full_name() or request.user.username,
		})

class DashboardRevenueChartAPIView(APIView):
	permission_classes = [permissions.IsAuthenticated]
	def get(self, request):
		period = request.GET.get('period', 'week')
		now = timezone.now()
		if period == 'month':
			days = 30
		elif period == 'year':
			days = 365
		else:
			days = 7
		labels = []
		revenue = []
		defects = []
		orders_count = []
		sales = []
		for i in range(days):
			day = now - timezone.timedelta(days=days - i - 1)
			day_orders = Order.objects.filter(created_at__date=day.date())
			# Доход за день — по позициям, относящимся к заказам этого дня
			day_income = OrderItem.objects.filter(order__in=day_orders).aggregate(total=Sum(F('product__price') * F('quantity')))['total'] or 0
			# Брак: сумма из OrderDefect + сумма defective_quantity из EmployeeTask за этот день
			day_order_defects = OrderDefect.objects.filter(date__date=day.date()).aggregate(total=Sum('quantity'))['total'] or 0
			day_employee_defects = EmployeeTask.objects.filter(created_at__date=day.date()).aggregate(total=Sum('defective_quantity'))['total'] or 0
			day_defects = day_order_defects + day_employee_defects
			day_orders_num = day_orders.count()
			# Продажи — сумма quantities по позициям заказов этого дня
			day_sales = OrderItem.objects.filter(order__in=day_orders).aggregate(total=Sum('quantity'))['total'] or 0
			labels.append(day.strftime('%d.%m'))
			revenue.append(day_income)
			defects.append(day_defects)
			orders_count.append(day_orders_num)
			sales.append(day_sales)
		return Response({
			'labels': labels,
			'revenue': revenue,
			'defects': defects,
			'orders_count': orders_count,
			'sales': sales
		})

class StageViewSet(viewsets.ReadOnlyModelViewSet):
	queryset = OrderStage.objects.select_related(
		'workshop', 
		'order',
		'order__client',
		'order_item',
		'order_item__product',
		'order_item__order'
	).all().order_by('deadline', 'sequence')
	serializer_class = OrderStageSerializer
	permission_classes = [permissions.IsAuthenticated]

	def get_queryset(self):
		qs = super().get_queryset()
		status_param = self.request.query_params.get('status')
		if status_param:
			qs = qs.filter(status=status_param)
		workshop_param = self.request.query_params.get('workshop')
		if workshop_param:
			qs = qs.filter(workshop_id=workshop_param)
		return qs



class PlansMasterView(TemplateView):
    template_name = 'plans_master.html'
class PlansMasterDetailView(View):
	def get(self, request, stage_id):
		stage = get_object_or_404(OrderStage, pk=stage_id)
		context = {
			'stage': stage,
			'order': stage.order,
			'workshop': stage.workshop,
		}
		return render(request, 'orders/plans_master_detail.html', context)


# ===== АДМИНИСТРАТОР ЗАЯВОК =====

@method_decorator(login_required, name='dispatch')
class AdminRequestsView(View):
    """Главная страница администратора заявок"""
    def get(self, request):
        from apps.finance.models import Request
        from apps.clients.models import Client
        
        # Определяем мобильное устройство
        user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
        is_mobile = any(m in user_agent for m in ['android', 'iphone', 'ipad', 'mobile'])
        
        # Получаем клиентов с заявками
        clients_with_requests = Client.objects.filter(
            requests__isnull=False
        ).annotate(
            requests_count=Count('requests'),
            last_request_date=Max('requests__created_at')
        ).filter(requests_count__gt=0).order_by('-last_request_date')
        
        # Статистика
        stats = {
            'pending': Request.objects.filter(status='pending').count(),
            'approved': Request.objects.filter(status='approved').count(),
            'in_production': Request.objects.filter(status='in_production').count(),
        }
        
        # Подготавливаем JSON для JavaScript
        clients_json = []
        for client in clients_with_requests:
            # Получаем количество заявок по статусам
            pending_count = client.requests.filter(status='pending').count()
            approved_count = client.requests.filter(status='approved').count()
            in_production_count = client.requests.filter(status='in_production').count()
            
            clients_json.append({
                'id': client.id,
                'name': client.name,
                'company': client.company or '',
                'phone': client.phone or '',
                'email': client.email or '',
                'requests_count': client.requests_count,
                'pending_count': pending_count,
                'approved_count': approved_count,
                'in_production_count': in_production_count,
                'last_request_date': client.last_request_date.strftime('%d.%m.%Y') if client.last_request_date else ''
            })
        
        context = {
            'clients_with_requests': clients_with_requests,
            'clients_with_requests_json': json.dumps(clients_json, ensure_ascii=False, default=str),
            'stats': stats,
        }
        
        # Выбираем шаблон в зависимости от устройства
        template = 'orders/admin_requests_mobile.html' if is_mobile else 'orders/admin_requests.html'
        return render(request, template, context)


class AdminClientRequestsView(View):
	"""Страница заявок конкретного клиента с канбан-доской"""
	def get(self, request, client_id):
		from apps.finance.models import Request
		from apps.clients.models import Client
		from apps.operations.workshops.models import Workshop
		
		# Определяем мобильное устройство
		user_agent = request.META.get('HTTP_USER_AGENT', '').lower()
		is_mobile = any(m in user_agent for m in ['android', 'iphone', 'ipad', 'mobile'])
		
		client = get_object_or_404(Client, pk=client_id)
		
		# Получаем все заявки клиента
		requests = Request.objects.filter(client=client).order_by('-created_at')
		
		# Получаем все цеха
		workshops = Workshop.objects.all().order_by('id')
		
		# Группируем заявки по статусам
		pending_requests = requests.filter(status='pending')
		approved_requests = requests.filter(status='approved')
		in_production_requests = requests.filter(status='in_production')
		
		context = {
			'client': client,
			'requests': requests,
			'workshops': workshops,
			'pending_requests': pending_requests,
			'approved_requests': approved_requests,
			'in_production_requests': in_production_requests,
		}
		
		# Выбираем шаблон в зависимости от устройства
		template = 'orders/admin_client_requests_mobile.html' if is_mobile else 'orders/admin_client_requests.html'
		return render(request, template, context)


class ApproveRequestAPIView(APIView):
	"""API для одобрения заявки и создания заказа"""
	permission_classes = [permissions.IsAuthenticated]
	
	def post(self, request, request_id):
		try:
			from apps.finance.models import Request
			request_obj = get_object_or_404(Request, pk=request_id)
			
			if request_obj.status != 'pending':
				return Response({
					'error': 'Заявка уже не в статусе ожидания'
				}, status=400)
			
			# Одобряем заявку и создаем заказ
			success, message = request_obj.approve_and_create_order(request.user)
			
			if success:
				return Response({
					'message': message,
					'order_id': request_obj.order.id if request_obj.order else None,
					'status': 'success'
				})
			else:
				return Response({
					'error': message
				}, status=400)
				
		except Exception as e:
			return Response({
				'error': f'Ошибка одобрения заявки: {str(e)}'
			}, status=500)


@method_decorator(login_required, name='dispatch')
class ExportRequestsExcelView(View):
	"""Экспорт заявок в Excel файл"""
	def get(self, request):
		from apps.finance.models import Request
		from apps.clients.models import Client
		
		# Получаем все заявки с детальной информацией
		requests = Request.objects.select_related('client').prefetch_related('items__product').all().order_by('-created_at')
		
		# Создаем Excel файл
		wb = Workbook()
		
		# Удаляем дефолтный лист
		wb.remove(wb.active)
		
		# Стили для заголовков
		header_font = Font(bold=True, size=12)
		header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		header_alignment = Alignment(horizontal="center", vertical="center")
		border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)
		
		# Создаем лист для каждой заявки
		for request_obj in requests:
			# Создаем лист с названием заявки (обрезаем до 31 символа)
			sheet_name = f"Заявка {request_obj.id}"
			if len(sheet_name) > 31:
				sheet_name = f"Заявка{request_obj.id}"
			
			ws = wb.create_sheet(title=sheet_name)
			
			# Заголовок заявки
			ws['A1'] = f"Заказ №{request_obj.id}-{request_obj.name} {request_obj.created_at.strftime('%d/%m/%y')}"
			ws['A1'].font = Font(bold=True, size=14)
			ws.merge_cells('A1:E1')
			
			# Информация о клиенте
			ws['A3'] = "Клиент:"
			ws['A3'].font = header_font
			ws['B3'] = request_obj.client.name
			ws['B3'].font = Font(size=12)
			
			ws['A4'] = "Компания:"
			ws['A4'].font = header_font
			ws['B4'] = request_obj.client.company or ""
			ws['B4'].font = Font(size=12)
			
			ws['A5'] = "Телефон:"
			ws['A5'].font = header_font
			ws['B5'] = request_obj.client.phone or ""
			ws['B5'].font = Font(size=12)
			
			# Заголовки таблицы
			headers = ['№', 'Материал', 'Размер', 'Шт', 'Операции']
			for col, header in enumerate(headers, 1):
				cell = ws.cell(row=7, column=col, value=header)
				cell.font = header_font
				cell.fill = header_fill
				cell.alignment = header_alignment
				cell.border = border
			
			# Данные товаров
			row = 8
			for idx, item in enumerate(request_obj.items.all(), 1):
				# Определяем материал
				material = item.product.name
				if item.glass_type:
					material += f" ({item.glass_type})"
				
				# Размер
				size = item.size or ""
				
				# Количество
				quantity = item.quantity
				
				# Операции (на основе типа продукта)
				operations = []
				if item.product.is_glass:
					operations.extend(['Распил', 'ЧПУ', 'Пескоструй', 'УФ печать'])
				else:
					operations.extend(['Распил', 'ЧПУ', 'Пресс', 'Кромка', 'Шлифовка', 'Грунтовка', 'Покраска'])
				
				operations_text = ", ".join(operations)
				
				# Записываем данные
				ws.cell(row=row, column=1, value=idx).border = border
				ws.cell(row=row, column=2, value=material).border = border
				ws.cell(row=row, column=3, value=size).border = border
				ws.cell(row=row, column=4, value=quantity).border = border
				ws.cell(row=row, column=5, value=operations_text).border = border
				
				row += 1
			
			# Итоговая строка
			total_quantity = sum(item.quantity for item in request_obj.items.all())
			ws.cell(row=row, column=1, value="Общий").font = Font(bold=True)
			ws.cell(row=row, column=1).border = border
			ws.cell(row=row, column=4, value=f"{total_quantity}шт").font = Font(bold=True)
			ws.cell(row=row, column=4).border = border
			
			# Настройка ширины столбцов
			ws.column_dimensions['A'].width = 8
			ws.column_dimensions['B'].width = 40
			ws.column_dimensions['C'].width = 15
			ws.column_dimensions['D'].width = 10
			ws.column_dimensions['E'].width = 50
			
			# Дополнительная информация
			row += 3
			ws.cell(row=row, column=1, value="Комментарий:").font = header_font
			ws.cell(row=row, column=2, value=request_obj.comment or "").font = Font(size=12)
			
			row += 1
			ws.cell(row=row, column=1, value="Общая сумма:").font = header_font
			ws.cell(row=row, column=2, value=f"{request_obj.total_amount or 0} сом").font = Font(size=12, bold=True)
			
			row += 1
			ws.cell(row=row, column=1, value="Статус:").font = header_font
			status_display = {
				'pending': 'Ожидает',
				'approved': 'Одобрена',
				'in_production': 'В производстве',
				'rejected': 'Отклонена'
			}.get(request_obj.status, request_obj.status)
			ws.cell(row=row, column=2, value=status_display).font = Font(size=12)
		
		# Создаем ответ
		response = HttpResponse(
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
		response['Content-Disposition'] = f'attachment; filename="заявки_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
		
		# Сохраняем файл
		wb.save(response)
		return response


@method_decorator(login_required, name='dispatch')
class ExportRequestsExcelForClientView(View):
	"""Экспорт заявок конкретного клиента в Excel файл с точным форматом как на фотографии"""
	def get(self, request, client_id):
		from apps.finance.models import Request
		from apps.clients.models import Client
		from apps.operations.workshops.models import Workshop
		
		# Получаем клиента
		client = get_object_or_404(Client, pk=client_id)
		
		# Получаем все заявки клиента с детальной информацией
		requests = Request.objects.filter(client=client).select_related('client').prefetch_related('items__product').order_by('-created_at')
		
		# Получаем цеха из БД (кроме ID 2)
		workshops = Workshop.objects.exclude(id=2).order_by('id')
		
		# Создаем Excel файл
		wb = Workbook()
		
		# Удаляем дефолтный лист
		wb.remove(wb.active)
		
		# Стили для заголовков
		header_font = Font(bold=True, size=10, color="FFFFFF")  # Уменьшаем размер с 12 до 10
		header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
		header_alignment = Alignment(horizontal="center", vertical="center")
		border = Border(
			left=Side(style='thin'),
			right=Side(style='thin'),
			top=Side(style='thin'),
			bottom=Side(style='thin')
		)
		
		# Создаем лист для каждой заявки
		for request_obj in requests:
			# Создаем лист с названием заявки (обрезаем до 31 символа)
			sheet_name = f"Заявка {request_obj.id}"
			if len(sheet_name) > 31:
				sheet_name = f"Заявка{request_obj.id}"
			
			ws = wb.create_sheet(title=sheet_name)
			
			# Заголовок заявки (точно как на фотографии)
			ws['A1'] = f"Заказ №{request_obj.id}-{request_obj.name} {request_obj.created_at.strftime('%d/%m/%y')}"
			ws['A1'].font = Font(bold=True, size=12)  # Уменьшаем размер шрифта с 14 до 12
			ws.merge_cells('A1:I1')  # Объединяем все 9 столбцов
			ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
			
			# Заголовки таблицы (точно как на фотографии)
			headers = ['№', 'цеха', 'материал', 'размер', 'шт', '', '', '', '']
			for col, header in enumerate(headers, 1):
				cell = ws.cell(row=2, column=col, value=header)
				cell.font = header_font
				cell.fill = header_fill
				cell.alignment = header_alignment
				cell.border = border
			
			# Данные товаров (точно как на фотографии)
			row = 3
			workshop_num = 1
			total_quantity = 0
			
			# Получаем товары из заявки
			items = list(request_obj.items.all())
			
			# Создаем данные для каждого цеха из БД
			for workshop in workshops:
				# Определяем количество строк для каждого цеха
				if workshop.id == 1:  # Распил
					rows_for_workshop = 3
				elif workshop.id == 2:  # ЧПУ (исключен)
					continue
				elif workshop.id == 3:  # Заготовка
					rows_for_workshop = 2
				elif workshop.id == 4:  # Пресс
					rows_for_workshop = 2
				elif workshop.id == 8:  # Грутовка
					rows_for_workshop = 2
				elif workshop.id == 9:  # Шкурка белый
					rows_for_workshop = 6
				elif workshop.id == 10:  # Покраска
					rows_for_workshop = 2
				elif workshop.id == 11:  # Упаковка
					rows_for_workshop = 2
				else:
					rows_for_workshop = 1
				
				# Для каждого товара создаем отдельные строки
				for item in items:
					for workshop_row in range(rows_for_workshop):
						# Первая строка цеха содержит номер и название
						if workshop_row == 0:
							ws.cell(row=row, column=1, value=workshop_num).border = border
							ws.cell(row=row, column=2, value=workshop.name).border = border
						else:
							# Последующие строки цеха пустые в первых двух столбцах
							ws.cell(row=row, column=1, value="").border = border
							ws.cell(row=row, column=2, value="").border = border
						
						# Заполняем данные товаров только для определенных цехов
						if workshop.id == 1 and workshop_row == 0:  # Распил
							material = item.product.name  # Только название товара без скобок
							size = item.size or "80-200"
							# До пресса x2 от количества заявки
							quantity = item.quantity * 2
							# Добавляем 1 к размеру до пресса
							size_with_plus_one = self.add_one_to_size(size)
							total_quantity += quantity
							
							ws.cell(row=row, column=3, value=material).border = border
							ws.cell(row=row, column=4, value=size_with_plus_one).border = border
							ws.cell(row=row, column=5, value=quantity).border = border
							ws.cell(row=row, column=6, value="").border = border
							
						elif workshop.id == 1 and workshop_row == 1:  # Распил вторая строка
							# Берем данные из БД
							material = f"{item.product.name} МДФ {item.size or '1,0'}"
							ws.cell(row=row, column=3, value=material).border = border
							ws.cell(row=row, column=4, value="").border = border
							ws.cell(row=row, column=5, value="").border = border
							ws.cell(row=row, column=6, value="").border = border
							
						elif workshop.id == 1 and workshop_row == 2:  # Распил третья строка
							# Берем данные из БД
							material = f"{item.product.name} стекло"
							size = item.size or "30 40"
							ws.cell(row=row, column=3, value=material).border = border
							ws.cell(row=row, column=4, value=size).border = border
							ws.cell(row=row, column=5, value="").border = border
							ws.cell(row=row, column=6, value="").border = border
							
						elif workshop.id == 3 and workshop_row == 0:  # Заготовка
							material = f"{item.product.name} ГЛУХОЙ"  # Убираем скобки с цветом
							size = item.size or "80-200"
							# До пресса x2 от количества заявки
							quantity = item.quantity * 2
							# Добавляем 1 к размеру до пресса
							size_with_plus_one = self.add_one_to_size(size)
							
							ws.cell(row=row, column=3, value=material).border = border
							ws.cell(row=row, column=4, value=size_with_plus_one).border = border
							ws.cell(row=row, column=5, value=quantity).border = border
							ws.cell(row=row, column=6, value="").border = border
							
						elif workshop.id == 4 and workshop_row == 0:  # Пресс
							material = f"{item.product.name} ГЛУХОЙ"  # Убираем скобки с цветом
							size = item.size or "80-200"
							# На прессе и после - реальное количество
							quantity = item.quantity
							
							ws.cell(row=row, column=3, value=material).border = border
							ws.cell(row=row, column=4, value=size).border = border
							ws.cell(row=row, column=5, value=quantity).border = border
							ws.cell(row=row, column=6, value="").border = border
							
						elif workshop.id == 8 and workshop_row == 0:  # Грутовка
							quantity = item.quantity  # Реальное количество из БД
							ws.cell(row=row, column=3, value="").border = border
							ws.cell(row=row, column=4, value="").border = border
							ws.cell(row=row, column=5, value=f"{quantity}шт").border = border
							ws.cell(row=row, column=6, value="").border = border
							
						elif workshop.id == 9 and workshop_row == 0:  # Шкурка белый
							quantity = item.quantity  # Реальное количество из БД
							ws.cell(row=row, column=3, value="").border = border
							ws.cell(row=row, column=4, value="").border = border
							ws.cell(row=row, column=5, value=f"{quantity}шт").border = border
							ws.cell(row=row, column=6, value="").border = border
							
						elif workshop.id == 10 and workshop_row == 0:  # Покраска
							material = f"{item.product.name} ГЛУХОЙ"  # Убираем скобки с цветом
							size = item.size or "80-200"
							# На прессе и после - реальное количество
							quantity = item.quantity
							
							ws.cell(row=row, column=3, value=material).border = border
							ws.cell(row=row, column=4, value=size).border = border
							ws.cell(row=row, column=5, value=quantity).border = border
							ws.cell(row=row, column=6, value="").border = border
						
						else:
							# Для всех остальных строк заполняем пустые ячейки с границами
							ws.cell(row=row, column=3, value="").border = border
							ws.cell(row=row, column=4, value="").border = border
							ws.cell(row=row, column=5, value="").border = border
							ws.cell(row=row, column=6, value="").border = border
						
						# Заполняем пустые столбцы 7-9 границами для всех строк
						for col in range(7, 10):
							ws.cell(row=row, column=col, value="").border = border
						
						row += 1
				
				workshop_num += 1
			
			# Итоговая строка (точно как на фотографии)
			ws.cell(row=row, column=1, value="общий").font = Font(bold=True)
			ws.cell(row=row, column=1).border = border
			ws.cell(row=row, column=1).alignment = Alignment(horizontal="right")
			
			# Вычисляем общее количество из БД
			total_from_db = sum(item.quantity for item in items) if items else 0
			ws.cell(row=row, column=7, value=f"{total_from_db}шт").font = Font(bold=True)  # В 7-м столбце как на фотографии
			ws.cell(row=row, column=7).border = border
			
			# Заполняем границы для итоговой строки
			for col in range(2, 10):
				ws.cell(row=row, column=col, value="").border = border
			
			# Настройка ширины столбцов для растягивания на всю ширину листа
			ws.column_dimensions['A'].width = 8   # №
			ws.column_dimensions['B'].width = 20  # цеха
			ws.column_dimensions['C'].width = 35  # материал - увеличиваем
			ws.column_dimensions['D'].width = 18  # размер - увеличиваем
			ws.column_dimensions['E'].width = 12  # шт - увеличиваем
			ws.column_dimensions['F'].width = 18  # 6-й столбец - увеличиваем
			ws.column_dimensions['G'].width = 15  # 7-й столбец - увеличиваем
			ws.column_dimensions['H'].width = 12  # 8-й столбец - увеличиваем
			ws.column_dimensions['I'].width = 12  # 9-й столбец - увеличиваем
			
			# Настройка ориентации страницы на альбомную
			ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
			ws.page_setup.fitToPage = True
			ws.page_setup.fitToHeight = 1
			ws.page_setup.fitToWidth = 1
			
			# Настройка отступов страницы для максимального использования пространства
			ws.page_margins.left = 0.3
			ws.page_margins.right = 0.3
			ws.page_margins.top = 0.3
			ws.page_margins.bottom = 0.3
			ws.page_margins.header = 0.2
			ws.page_margins.footer = 0.2
		
		# Создаем ответ
		response = HttpResponse(
			content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
		)
		response['Content-Disposition'] = f'attachment; filename="заявки_{client.name}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
		
		# Сохраняем файл
		wb.save(response)
		return response
	
	def add_one_to_size(self, size):
		"""Добавляет 1 к размеру до пресса (например, 80-200 -> 81-201)"""
		if not size:
			return size
		
		try:
			# Разбиваем размер по дефису
			parts = size.split('-')
			if len(parts) == 2:
				first_part = int(parts[0]) + 1
				second_part = int(parts[1]) + 1
				return f"{first_part}-{second_part}"
			else:
				# Если размер не в формате X-Y, возвращаем как есть
				return size
		except (ValueError, TypeError):
			# Если не удается преобразовать в числа, возвращаем как есть
			return size